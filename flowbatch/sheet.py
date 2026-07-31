"""Очередь из Excel-файла (формат UiPath-очереди ZASTRYALO).

Лист IMG_QUEUE и VID_QUEUE читаются в те же Job, что и jobs.yaml, а статус
пишется обратно в книгу: STATUS, RESULT_PATH, TIMESTAMP, ATTEMPTS. Так таблица
остаётся единственным источником правды и её видно глазами.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .queue import Job

# Статусы в колонке 03_STATUS.
ST_TODO = "TODO"
ST_IN_PROGRESS = "IN_PROGRESS"
ST_DONE = "DONE"
ST_ERROR = "ERROR"
ST_SKIP = "SKIP"

SHEET_IMG = "IMG_QUEUE"
SHEET_VID = "VID_QUEUE"
SHEET_SETTINGS = "SETTINGS"

# Имена колонок различаются между листами, потому что нумерация в префиксе
# отражает позицию, а у видео колонок меньше. Держим карту явно.
COLS_IMG = {
    "id": "01_ID",
    "order": "02_ORDER",
    "status": "03_STATUS",
    "batch": "04_BATCH",
    "shot_name": "08_SHOT_NAME",
    "aspect": "10_ASPECT",
    # Исходные колонки референсов. Читаем именно их, а не 16_REF_LIST:
    # тот — формула TEXTJOIN, и её вычисленное значение живёт только в кэше,
    # который любая перезапись книги обнуляет.
    "ref_1": "11_REF_HEROES",
    "ref_2": "12_REF_TONGUE",
    "ref_3": "13_REF_BONE",
    "ref_4": "14_REF_PRODUCT",
    "ref_list": "16_REF_LIST",
    "prompt": "17_PROMPT",
    "prompt_1line": "18_PROMPT_1LINE",
    "output_name": "19_OUTPUT_NAME",
    "result_path": "20_RESULT_PATH",
    "attempts": "21_ATTEMPTS",
    "timestamp": "22_TIMESTAMP",
    "notes": "23_NOTES",
}

COLS_VID = {
    "id": "01_ID",
    "order": "02_ORDER",
    "status": "03_STATUS",
    "batch": "04_BATCH",
    "shot_name": "08_SHOT_NAME",
    "aspect": "10_ASPECT",
    "duration": "11_DURATION_SEC",
    "source_image": "12_SOURCE_IMAGE",
    "ref_2": "13_REF_PRODUCT",
    "ref_list": "14_REF_LIST",
    "prompt": "15_PROMPT",
    "prompt_1line": "16_PROMPT_1LINE",
    "output_name": "17_OUTPUT_NAME",
    "result_path": "18_RESULT_PATH",
    "attempts": "19_ATTEMPTS",
    "timestamp": "20_TIMESTAMP",
    "notes": "21_NOTES",
}


@dataclass
class SheetRow:
    """Строка очереди вместе с координатами для записи результата обратно."""

    job: Job
    sheet: str
    row_idx: int  # 1-based номер строки в книге
    order: int
    batch: str
    status: str
    output_name: str
    shot_name: str


class SheetQueue:
    """Чтение очереди из книги Excel и запись результатов обратно."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Не найден файл очереди: {self.path.resolve()}")
        self._rows: list[SheetRow] = []
        self.settings: dict[str, str] = {}
        self._frozen = False
        # Ячейки, чьи формулы пришлось заморозить в значения (для отчёта).
        self.frozen_cells: list[str] = []

    # ------------------------------------------------------------------ чтение

    def load(self, statuses: Iterable[str] = (ST_TODO,)) -> list[SheetRow]:
        """Прочитать обе очереди. По умолчанию берутся только строки TODO."""
        wanted = {s.upper() for s in statuses}
        wb = load_workbook(self.path, data_only=True)
        self.settings = self._read_settings(wb)

        rows: list[SheetRow] = []
        if SHEET_IMG in wb.sheetnames:
            rows += self._read_queue(wb[SHEET_IMG], SHEET_IMG, COLS_IMG, "image", wanted)
        if SHEET_VID in wb.sheetnames:
            rows += self._read_queue(wb[SHEET_VID], SHEET_VID, COLS_VID, "video", wanted)
        wb.close()

        # 02_ORDER задаёт порядок внутри листа; картинки идут раньше видео,
        # потому что видео ссылаются на их результаты.
        prio = {SHEET_IMG: 0, SHEET_VID: 1}
        rows.sort(key=lambda r: (prio.get(r.sheet, 9), r.order))
        self._rows = rows
        return rows

    @property
    def project_name(self) -> str | None:
        """Имя проекта Flow из SETTINGS (ключ PROJECT_NAME), если задано."""
        return self.settings.get("PROJECT_NAME") or None

    @staticmethod
    def _read_settings(wb: Any) -> dict[str, str]:
        if SHEET_SETTINGS not in wb.sheetnames:
            return {}
        out: dict[str, str] = {}
        for row in wb[SHEET_SETTINGS].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()
        return out

    def _read_queue(
        self,
        ws: Any,
        sheet_name: str,
        cols: dict[str, str],
        kind: str,
        wanted: set[str],
    ) -> list[SheetRow]:
        header = [c.value for c in ws[1]]
        idx = {name: i for i, name in enumerate(header) if name}

        missing = [c for key, c in cols.items() if c not in idx and key in ("id", "status", "prompt")]
        if missing:
            raise ValueError(f"{sheet_name}: нет обязательных колонок {missing}")

        def get(row: tuple[Any, ...], key: str) -> Any:
            col = cols.get(key)
            if col is None or col not in idx:
                return None
            return row[idx[col]]

        out: list[SheetRow] = []
        for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            job_id = get(row, "id")
            if job_id is None or str(job_id).strip() == "":
                continue
            status = str(get(row, "status") or "").strip().upper()
            if wanted and status not in wanted:
                continue

            prompt = get(row, "prompt")
            if prompt is None or not str(prompt).strip():
                # запасной вариант — однострочная версия
                prompt = get(row, "prompt_1line")
            prompt = str(prompt or "").strip()
            if not prompt:
                raise ValueError(f"{sheet_name} строка {n} ({job_id}): пустой промпт")

            refs = self._collect_refs(row, get, kind)
            duration = get(row, "duration")
            duration = int(duration) if duration not in (None, "") else None

            out.append(
                SheetRow(
                    job=Job(
                        id=str(job_id).strip(),
                        kind=kind,  # type: ignore[arg-type]
                        prompt=prompt,
                        refs=refs,
                        duration=duration if kind == "video" else None,
                        output_name=str(get(row, "output_name") or "").strip() or None,
                    ),
                    sheet=sheet_name,
                    row_idx=n,
                    order=int(get(row, "order") or 0),
                    batch=str(get(row, "batch") or "").strip(),
                    status=status,
                    output_name=str(get(row, "output_name") or "").strip(),
                    shot_name=str(get(row, "shot_name") or "").strip(),
                )
            )
        return out

    @staticmethod
    def _collect_refs(row: tuple[Any, ...], get: Any, kind: str) -> list[str]:
        """Собрать референсы из ИСХОДНЫХ колонок, а не из формулы REF_LIST.

        16_REF_LIST — это =TEXTJOIN(";",TRUE,K:N), то есть склейка тех же
        колонок. Её значение хранится только в кэше книги, и любая перезапись
        файла (в том числе нашей же записью статуса) этот кэш обнуляет.
        Поэтому основной источник — колонки REF_*, а REF_LIST лишь дополняет:
        мало ли, кто-то заполнил её руками.
        """
        refs: list[str] = []
        if kind == "video":
            src = get(row, "source_image")
            if src and str(src).strip():
                refs.append(str(src).strip())

        for key in ("ref_1", "ref_2", "ref_3", "ref_4"):
            val = get(row, key)
            if val and str(val).strip():
                refs.append(str(val).strip())

        raw = get(row, "ref_list")
        if raw and str(raw).strip():
            refs += [p.strip() for p in str(raw).split(";") if p.strip()]

        # убираем дубли, сохраняя порядок
        seen: set[str] = set()
        return [r for r in refs if not (r in seen or seen.add(r))]

    # ------------------------------------------------------------------ запись

    def _freeze_formulas(self, wb: Any) -> None:
        """Заменить формулы их вычисленными значениями — один раз за сессию.

        openpyxl пишет формулы без кэша результата, поэтому первая же наша
        запись статуса обнулила бы всё, что читается по кэшу: 16_REF_LIST,
        15_REF_COUNT и любой другой производный столбец. Для UiPath, читающего
        книгу через Read Range Workbook, они стали бы пустыми.

        Замораживание сохраняет ДАННЫЕ ценой автопересчёта: значение остаётся
        ровно тем, что видно в файле сейчас. Если после этого править колонки
        REF_* руками, REF_LIST сам не обновится — пересобирать очередь надо из
        исходного файла.
        """
        if self._frozen:
            return
        self._frozen = True
        try:
            cached = load_workbook(self.path, data_only=True)
        except Exception:  # noqa: BLE001 — не смогли прочитать кэш, ничего не трогаем
            return

        for ws in wb.worksheets:
            if ws.title not in cached.sheetnames:
                continue
            wsv = cached[ws.title]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        val = wsv.cell(row=cell.row, column=cell.column).value
                        if val is None:
                            # Кэш пуст (формула с пустым результатом или уже
                            # убитый кэш) — замена дала бы None вместо формулы,
                            # то есть потерю и данных, и определения. Не трогаем.
                            continue
                        cell.value = val
                        self.frozen_cells.append(f"{ws.title}!{cell.coordinate}")
        cached.close()

    def write_result(
        self,
        row: SheetRow,
        status: str,
        result_path: str | None = None,
        bump_attempts: bool = False,
        note: str | None = None,
    ) -> None:
        """Записать статус и результат обратно в книгу.

        Открываем и сохраняем на каждую строку: прогон длинный, и процесс могут
        убить в любой момент — тогда всё, что успело записаться, останется.
        """
        cols = COLS_IMG if row.sheet == SHEET_IMG else COLS_VID
        wb = load_workbook(self.path)
        self._freeze_formulas(wb)
        ws = wb[row.sheet]
        header = [c.value for c in ws[1]]
        idx = {name: i + 1 for i, name in enumerate(header) if name}

        def put(key: str, value: Any) -> None:
            col = cols.get(key)
            if col and col in idx:
                ws.cell(row=row.row_idx, column=idx[col]).value = value

        put("status", status)
        put("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if result_path is not None:
            put("result_path", str(result_path))
        if bump_attempts:
            col = cols.get("attempts")
            if col and col in idx:
                cell = ws.cell(row=row.row_idx, column=idx[col])
                try:
                    cell.value = int(cell.value or 0) + 1
                except (TypeError, ValueError):
                    cell.value = 1
        if note:
            col = cols.get("notes")
            if col and col in idx:
                cell = ws.cell(row=row.row_idx, column=idx[col])
                prev = str(cell.value or "").strip()
                cell.value = (prev + " | " if prev else "") + note[:250]

        wb.save(self.path)
        wb.close()


def filter_rows(
    rows: list[SheetRow],
    only: str | None = None,
    from_id: str | None = None,
    limit: int | None = None,
    batch: str | None = None,
    kind: str | None = None,
) -> list[SheetRow]:
    """Фильтры очереди: --only / --from / --batch / --kind / --limit."""
    result = rows
    if kind:
        result = [r for r in result if r.job.kind == kind]
    if batch:
        result = [r for r in result if r.batch.upper() == batch.upper()]
    if only:
        result = [r for r in result if r.job.id == only]
        if not result:
            raise ValueError(f"--only {only!r}: такой строки в очереди нет")
    if from_id:
        ids = [r.job.id for r in result]
        if from_id not in ids:
            raise ValueError(f"--from {from_id!r}: такой строки в очереди нет")
        result = result[ids.index(from_id):]
    if limit is not None:
        if limit < 1:
            raise ValueError("--limit должен быть >= 1")
        result = result[:limit]
    return result
